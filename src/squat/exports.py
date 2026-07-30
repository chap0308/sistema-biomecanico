"""Research-ready Excel and PDF exports for squat comparisons."""

from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any, Iterable

from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.squat.comparison import (
    CaseComparison,
    DatasetPerformance,
    PatternComparison,
)

_PATTERN_NAMES = {
    "trunk_lateral_inclination": "Inclinación lateral del tronco",
    "pelvis_lateral_shift": "Desplazamiento lateral de pelvis",
    "visible_dynamic_valgus": "Valgo dinámico visible",
    "bilateral_asymmetry": "Asimetría bilateral observable",
}

_VALUE_LABELS = {
    "anterior": "Vista anterior",
    "frontal": "Plano frontal",
    "sin_carga_externa": "Sin carga externa",
    "no_etiquetado": "No etiquetado",
    "ninguno_de_los_anteriores": "Ninguno de los anteriores",
    "masculino": "Masculino",
    "femenino": "Femenino",
    "adecuada": "Adecuada",
    "regular": "Regular",
    "deficiente": "Deficiente",
    "adecuado": "Adecuado",
    "no_verificada": "No verificada",
    "no_verificado": "No verificado",
    "completa": "Completa",
    "parcial_utilizable": "Parcial utilizable",
    "insuficiente": "Insuficiente",
    "ninguna": "Ninguna",
    "leve": "Leve",
    "moderada": "Moderada",
    "severa": "Severa",
    "plana": "Plana",
    "no_verificable": "No verificable",
    "continuo": "Continuo",
    "si": "Sí",
    "no": "No",
    "analisis_completo": "Análisis completo",
    "apto_para_analisis": "Apto para análisis",
    "revision_requerida": "Revisión requerida",
    "no_apto_para_analisis": "No apto para análisis",
    "presente": "Presente",
    "presente_izquierda": "Presente, izquierda",
    "presente_derecha": "Presente, derecha",
    "presente_bilateral": "Presente, bilateral",
    "presente_sin_direccion": "Presente, sin dirección",
    "ausente": "Ausente",
    "no_concluyente": "No concluyente",
    "izquierda": "Izquierda",
    "derecha": "Derecha",
    "bilateral": "Bilateral",
    "sin_direccion": "Sin dirección",
    "pct_ancho_hombros": "% del ancho inicial de hombros",
    "deg": "Grados",
    "consenso_guiado": "Consenso guiado",
    "coincidencia_directa": "Coincidencia directa",
    "mayoria_absoluta": "Mayoría absoluta",
    "visible_estable": "Visible y estable",
    "intermitente": "Intermitente",
    "no_disponible": "No disponible",
    "inclinacion_lateral_tronco": "Inclinación lateral del tronco",
    "desplazamiento_lateral_pelvis": "Desplazamiento lateral de pelvis",
    "valgo_dinamico_visible": "Valgo dinámico visible",
    "asimetria_bilateral_observable": "Asimetría bilateral observable",
    "trunk_lateral_inclination": "Inclinación lateral del tronco",
    "pelvis_lateral_shift": "Desplazamiento lateral de pelvis",
    "visible_dynamic_valgus": "Valgo dinámico visible",
    "bilateral_asymmetry": "Asimetría bilateral observable",
    "reposo": "Reposo",
    "descenso": "Descenso",
    "maxima_profundidad": "Máxima profundidad",
    "ascenso": "Ascenso",
    "cierre": "Cierre",
    "consolidada": "Consolidada",
    "consenso_requerido": "Consenso requerido",
    "evaluaciones_pendientes": "Evaluaciones pendientes",
    "general": "General",
}

_LANDMARK_LABELS = {
    "shoulder": "Hombro",
    "hip": "Cadera",
    "knee": "Rodilla",
    "ankle": "Tobillo",
    "heel": "Talón",
    "foot_index": "Punta del pie",
    "nose": "Nariz",
}

_TECHNICAL_SHEETS = {
    "landmarks.csv": "Puntos anatómicos",
    "frame_quality.csv": "Calidad fotogramas",
    "frame_phases.csv": "Fases",
    "repetitions.csv": "Repeticiones",
    "biomechanical_frame_metrics.csv": "Biomecánica fotogramas",
    "biomechanical_repetition_metrics.csv": "Biomecánica repeticiones",
    "rule_evidence.csv": "Evidencia reglas",
}

_TECHNICAL_HEADERS = {
    "frame_index": "N.° de fotograma",
    "timestamp_seconds": "Tiempo (s)",
    "landmark": "Punto anatómico clave",
    "x": "Coordenada X normalizada",
    "y": "Coordenada Y normalizada",
    "z": "Profundidad relativa Z",
    "visibility": "Visibilidad",
    "presence": "Presencia",
    "pose_detected": "Pose detectada",
    "valid_for_analysis": "Válido para análisis",
    "detected_keypoints": "Puntos anatómicos clave detectados",
    "minimum_critical_visibility": "Visibilidad crítica mínima",
    "invalid_reason": "Motivo de invalidez",
    "hip_midpoint_y": "Centro de caderas Y",
    "hip_midpoint_y_smoothed": "Centro de caderas Y suavizado",
    "repetition_index": "N.° de repetición",
    "phase": "Fase",
    "start_frame": "Fotograma inicial",
    "peak_depth_frame": "Fotograma de máxima profundidad",
    "end_frame": "Fotograma final",
    "start_seconds": "Inicio (s)",
    "peak_depth_seconds": "Máxima profundidad (s)",
    "end_seconds": "Final (s)",
    "descent_duration_seconds": "Duración del descenso (s)",
    "ascent_duration_seconds": "Duración del ascenso (s)",
    "total_duration_seconds": "Duración total (s)",
    "peak_hip_midpoint_y": "Centro de caderas Y en máxima profundidad",
    "valid_frames_percentage": "Fotogramas válidos (%)",
    "trunk_inclination_deg": "Inclinación lateral del tronco (°)",
    "pelvis_lateral_shift_pct": "Desplazamiento lateral de pelvis (%)",
    "left_knee_medial_deviation_pct": "Desviación medial rodilla izquierda (%)",
    "right_knee_medial_deviation_pct": "Desviación medial rodilla derecha (%)",
    "bilateral_alignment_difference_pct": "Diferencia de alineación bilateral (%)",
    "trunk_inclination_at_peak_deg": "Tronco en máxima profundidad (°)",
    "trunk_max_abs_deg": "Máxima inclinación absoluta del tronco (°)",
    "trunk_max_abs_frame": "Fotograma de máxima inclinación del tronco",
    "pelvis_shift_at_peak_pct": "Pelvis en máxima profundidad (%)",
    "pelvis_max_abs_shift_pct": "Máximo desplazamiento absoluto de pelvis (%)",
    "pelvis_max_abs_frame": "Fotograma de máximo desplazamiento de pelvis",
    "left_knee_medial_deviation_at_peak_pct": "Rodilla izquierda en máxima profundidad (%)",
    "right_knee_medial_deviation_at_peak_pct": "Rodilla derecha en máxima profundidad (%)",
    "left_knee_max_medial_deviation_pct": "Máxima desviación rodilla izquierda (%)",
    "right_knee_max_medial_deviation_pct": "Máxima desviación rodilla derecha (%)",
    "bilateral_alignment_difference_at_peak_pct": "Asimetría en máxima profundidad (%)",
    "bilateral_max_alignment_difference_pct": "Máxima asimetría bilateral (%)",
    "finding": "Compensación observable",
    "status": "Clasificación",
    "direction": "Dirección",
    "metric": "Métrica aplicada",
    "unit": "Unidad",
    "aggregate_value": "Valor calculado",
    "repetition_values": "Valores de la repetición",
    "repetition_states": "Estados de la repetición",
    "absent_max": "Umbral máximo de ausencia",
    "present_min": "Umbral mínimo de presencia",
    "ruleset_rationale": "Justificación de la regla",
}


def build_case_excel(
    *,
    case_record: dict[str, Any],
    case_report: dict[str, Any],
    comparison: CaseComparison,
    performance: DatasetPerformance,
    landmark_visibility: list[dict[str, Any]] | None = None,
) -> bytes:
    """Create one workbook containing Instruments 1-3 and derived analysis."""
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    _instrument_1_sheet(
        workbook,
        case_record,
        case_report,
        landmark_visibility or [],
    )
    _instrument_2_sheet(workbook, case_report)
    _instrument_3_sheet(workbook, comparison)
    _analysis_sheet(workbook, comparison.patterns)
    _metrics_sheet(workbook, performance, comparison)
    for sheet in workbook.worksheets:
        _format_sheet(sheet)
    _append_landmark_code_legend(workbook["Instrumento 1"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_technical_data_excel(*, artifacts: dict[str, bytes]) -> bytes:
    """Create readable tables without modifying canonical pipeline CSV files."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for filename, sheet_name in _TECHNICAL_SHEETS.items():
        content = artifacts.get(filename)
        if content is None:
            continue
        rows = list(csv.reader(StringIO(content.decode("utf-8-sig"))))
        if not rows:
            continue
        sheet = workbook.create_sheet(sheet_name)
        sheet.append([_TECHNICAL_HEADERS.get(value, _label(value)) for value in rows[0]])
        for row in rows[1:]:
            sheet.append([_technical_cell(value) for value in row])
        _format_sheet(sheet)
        sheet.freeze_panes = "A2"
    if not workbook.worksheets:
        sheet = workbook.create_sheet("Datos técnicos")
        sheet.append(["Estado"])
        sheet.append(["No se encontraron archivos CSV para este caso."])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_case_pdf(
    *,
    case_report: dict[str, Any],
    comparison: CaseComparison,
    performance: DatasetPerformance,
) -> bytes:
    """Create a concise investigator report without clinical diagnoses."""
    output = BytesIO()
    with PdfPages(output) as pdf:
        table_rows = [
            [
                f"R{row.repetition_index} · {_PATTERN_NAMES[row.pattern_key]}",
                _label(row.reference.label if row.reference else None),
                _label(row.system_label),
                _match_label(row.exact_match),
            ]
            for row in comparison.patterns
        ]
        chunks = [
            table_rows[index : index + 8]
            for index in range(0, len(table_rows), 8)
        ] or [[]]
        for page_index, rows in enumerate(chunks, start=1):
            figure = _pdf_figure_header(
                case_report=case_report,
                comparison=comparison,
                subtitle=f"Comparación por patrón · página {page_index}",
            )
            axis = figure.add_axes((0.07, 0.16, 0.86, 0.65))
            axis.axis("off")
            table = axis.table(
                cellText=rows,
                colLabels=[
                    "Patrón",
                    "Referencia experta",
                    "Sistema",
                    "Coincidencia",
                ],
                cellLoc="left",
                colLoc="left",
                loc="upper left",
                colWidths=[0.36, 0.24, 0.24, 0.16],
            )
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1, 1.8)
            pdf.savefig(figure, bbox_inches="tight")
            plt.close(figure)

        figure = _pdf_figure_header(
            case_report=case_report,
            comparison=comparison,
            subtitle="Desempeño técnico acumulado",
        )
        metrics = performance.overall
        metric_text = (
            "Desempeño acumulado del conjunto\n"
            f"Pares incluidos: {metrics.included_pairs}\n"
            f"Exactitud: {_percent(metrics.accuracy)}\n"
            f"Precisión: {_percent(metrics.precision)}\n"
            f"Sensibilidad: {_percent(metrics.sensitivity)}\n"
            f"Especificidad: {_percent(metrics.specificity)}\n"
            f"F1-score: {_decimal(metrics.f1_score)}\n"
            f"Kappa de Cohen: {_decimal(metrics.cohen_kappa)}\n"
            f"Kappa de Fleiss (tres expertos): "
            f"{_decimal(comparison.expert_fleiss_kappa)}\n"
            f"Ítems incluidos en Fleiss: {comparison.fleiss_items}\n"
            f"Pares no concluyentes excluidos: "
            f"{metrics.excluded_inconclusive_pairs}"
        )
        figure.text(
            0.08,
            0.72,
            metric_text,
            fontsize=10,
            linespacing=1.45,
            color="#1e293b",
        )
        figure.text(
            0.08,
            0.28,
            (
                "Este reporte describe compensaciones observables durante la "
                "sentadilla y no constituye un diagnóstico clínico. "
                "Los pares no concluyentes no se incluyen en las métricas."
            ),
            fontsize=8,
            color="#64748b",
            wrap=True,
        )
        pdf.savefig(figure, bbox_inches="tight")
        plt.close(figure)
    return output.getvalue()


def _pdf_figure_header(
    *,
    case_report: dict[str, Any],
    comparison: CaseComparison,
    subtitle: str,
) -> plt.Figure:
    figure = plt.figure(figsize=(8.27, 11.69))
    figure.patch.set_facecolor("#f8f5ec")
    figure.text(
        0.08,
        0.93,
        "Evaluación de sentadilla bilateral",
        fontsize=20,
        fontweight="bold",
        color="#123b42",
    )
    figure.text(
        0.08,
        0.89,
        f"Caso: {comparison.case_id}",
        fontsize=11,
        color="#334155",
    )
    figure.text(
        0.08,
        0.855,
        (
            f"{subtitle} | Pipeline: "
            f"{case_report.get('pipeline_version', 'N/D')} | "
            f"Evaluaciones enviadas: {comparison.submitted_evaluations}"
        ),
        fontsize=9,
        color="#64748b",
    )
    return figure


def _instrument_1_sheet(
    workbook: Workbook,
    record: dict[str, Any],
    report: dict[str, Any],
    landmark_visibility: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("Instrumento 1")
    sheet.append(["Campo", "Valor registrado"])
    registration = record.get("registration") or {}
    case = registration.get("case") or {}
    video = registration.get("video") or {}
    review = record.get("manual_protocol_review") or {}
    resolution = (
        f"{video.get('width_px')} × {video.get('height_px')} px"
        if video.get("width_px") and video.get("height_px")
        else None
    )
    fields = [
        ("Código del video", case.get("case_id")),
        ("Código del participante", case.get("participant_code")),
        ("Edad del participante", case.get("participant_age")),
        ("Sexo del participante", case.get("participant_sex")),
        ("Fecha y hora de creación", record.get("created_at")),
        ("Fecha del registro", review.get("record_date")),
        ("Fuente del video", review.get("video_source")),
        ("Ruta del video", video.get("path")),
        ("Vista de captura", case.get("view")),
        ("Plano de captura", case.get("plane")),
        ("Condición de carga", case.get("load_condition")),
        ("Dispositivo de captura", review.get("capture_device")),
        ("Resolución", resolution),
        ("Frecuencia de video", _with_unit(video.get("fps"), "fps")),
        (
            "Duración",
            _with_unit(video.get("duration_seconds"), "s"),
        ),
        ("Iluminación", review.get("lighting")),
        ("Fondo visual", review.get("background")),
        ("Visibilidad corporal", review.get("body_visibility")),
        ("Oclusiones", review.get("occlusions")),
        (
            "Sentadilla completa observable",
            review.get("complete_squat_observable"),
        ),
        ("Superficie", review.get("surface")),
        (
            "Soporte externo debajo de los talones",
            review.get("external_heel_support"),
        ),
        (
            "Contacto aparente de los talones",
            review.get("apparent_heel_contact"),
        ),
        (
            "Condición de apoyo conforme al protocolo",
            review.get("support_condition_compliant"),
        ),
        (
            "Observación del apoyo plantar",
            review.get("plantar_support_observation"),
        ),
        (
            "Video válido para procesamiento",
            registration.get("ready_for_pose"),
        ),
        ("Estado del análisis", report.get("status")),
    ]
    for label, value in fields:
        sheet.append([label, _display_value(value)])

    if landmark_visibility:
        sheet.append([])
        section_row = sheet.max_row + 1
        sheet.append(
            [
                "Disponibilidad computacional por repetición",
                "Promedio y cobertura derivados del análisis",
            ]
        )
        table_header_row = sheet.max_row + 1
        sheet.append(
            [
                "Repetición",
                "Grupo anatómico",
                "Promedio izquierdo o central",
                "Cobertura izquierda o central",
                "Estado izquierdo o central",
                "Promedio derecho",
                "Cobertura derecha",
                "Estado derecho",
                "Código de disponibilidad",
            ]
        )
        for row in _paired_landmark_rows(landmark_visibility):
            sheet.append(row)
        section_fill = PatternFill("solid", fgColor="DCE9E7")
        header_fill = PatternFill("solid", fgColor="315D63")
        for cell in sheet[section_row]:
            cell.font = Font(bold=True, color="123B42")
            cell.fill = section_fill
        for cell in sheet[table_header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill


def _instrument_2_sheet(workbook: Workbook, report: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Instrumento 2")
    pose = report.get("pose") or {}
    findings = report.get("findings") or {}
    decisions = findings.get("decisions") or []
    headers = [
        "Código del video",
        "Estado",
        "Fotogramas totales",
        "Fotogramas procesados correctamente",
        "Fotogramas válidos",
        "% fotogramas válidos",
        "% procesados correctamente",
        "Promedio de puntos anatómicos clave detectados por fotograma",
        "Reglas implementadas",
        "Compensaciones detectadas",
        "Versión de reglas",
    ]
    sheet.append(headers)
    sheet.append(
        [
            report.get("case_id"),
            _display_value(report.get("status")),
            pose.get("total_frames"),
            pose.get("processed_frames"),
            pose.get("valid_frames"),
            pose.get("valid_frames_percentage"),
            pose.get("processed_frames_percentage"),
            pose.get("mean_detected_keypoints"),
            len(decisions),
            _detected_findings_label(findings.get("detected_findings")),
            findings.get("ruleset_version"),
        ]
    )
    sheet.append([])
    sheet.append(
        [
            "Patrón",
            "Repetición",
            "Estado del sistema",
            "Dirección",
            "Valor agregado",
            "Unidad",
            "Umbral ausente",
            "Umbral presente",
        ]
    )
    for decision in decisions:
        sheet.append(
            [
                _display_value(decision.get("finding")),
                decision.get("repetition_index"),
                _display_value(decision.get("status")),
                _display_value(decision.get("direction")),
                decision.get("aggregate_value"),
                _display_value(decision.get("unit")),
                decision.get("absent_max"),
                decision.get("present_min"),
            ]
        )


def _instrument_3_sheet(
    workbook: Workbook,
    comparison: CaseComparison,
) -> None:
    sheet = workbook.create_sheet("Instrumento 3")
    sheet.append(
        [
            "Código del video",
            "Ejecución",
            "Patrón",
            "Evaluador 1",
            "Evaluador 2",
            "Evaluador 3",
            "Sistema",
            "Referencia final",
            "Método de consolidación",
        ]
    )
    for row in comparison.patterns:
        expert_labels = [
            _judgment_label(
                judgment.classification,
                judgment.observed_side,
                row.pattern_key,
            )
            for judgment in row.expert_judgments
        ]
        expert_labels.extend([""] * (3 - len(expert_labels)))
        sheet.append(
            [
                comparison.case_id,
                f"{comparison.case_id} - Repetición {row.repetition_index}",
                _PATTERN_NAMES[row.pattern_key],
                *expert_labels[:3],
                _label(row.system_label),
                _label(row.reference.label if row.reference else None),
                (
                    _display_value(row.reference.method)
                    if row.reference
                    else _display_value(row.reference_status)
                ),
            ]
        )


def _analysis_sheet(
    workbook: Workbook,
    rows: Iterable[PatternComparison],
) -> None:
    sheet = workbook.create_sheet("Matriz de análisis")
    sheet.append(
        [
            "Ejecución",
            "Patrón",
            "Referencia final",
            "Salida del sistema",
            "Coincidencia exacta",
            "Resultado binario",
            "Observación",
        ]
    )
    for row in rows:
        sheet.append(
            [
                f"Repetición {row.repetition_index}",
                _PATTERN_NAMES[row.pattern_key],
                _label(row.reference.label if row.reference else None),
                _label(row.system_label),
                _match_label(row.exact_match),
                row.binary_outcome or "No calculable",
                (
                    "Par excluido por estado no concluyente"
                    if row.reference
                    and (
                        row.reference.classification == "no_concluyente"
                        or row.system_classification == "no_concluyente"
                    )
                    else _display_value(row.reference_status)
                ),
            ]
        )


def _metrics_sheet(
    workbook: Workbook,
    performance: DatasetPerformance,
    comparison: CaseComparison,
) -> None:
    sheet = workbook.create_sheet("Métricas")
    sheet.append(
        [
            "Ámbito",
            "Pares incluidos",
            "Excluidos",
            "VP",
            "VN",
            "FP",
            "FN",
            "Exactitud",
            "Precisión",
            "Sensibilidad",
            "Especificidad",
            "F1-score",
            "Acuerdo exacto",
            "Kappa de Cohen",
            "Kappa de Fleiss (3 expertos)",
            "Ítems incluidos en Fleiss",
        ]
    )
    for metric in [performance.overall, *performance.by_pattern]:
        is_overall = metric.scope == "general"
        sheet.append(
            [
                _display_value(metric.scope),
                metric.included_pairs,
                metric.excluded_inconclusive_pairs,
                metric.true_positive,
                metric.true_negative,
                metric.false_positive,
                metric.false_negative,
                metric.accuracy,
                metric.precision,
                metric.sensitivity,
                metric.specificity,
                metric.f1_score,
                metric.exact_agreement,
                metric.cohen_kappa,
                comparison.expert_fleiss_kappa if is_overall else None,
                comparison.fleiss_items if is_overall else None,
            ]
        )


def _format_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="123B42")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        width = max(
            len(str(sheet.cell(row=row, column=column).value or ""))
            for row in range(1, sheet.max_row + 1)
        )
        sheet.column_dimensions[get_column_letter(column)].width = min(
            max(width + 2, 12),
            42,
        )
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _append_landmark_code_legend(sheet: Any) -> None:
    sheet.append([])
    sheet.append([])
    sheet.append(["Codificación de puntos anatómicos clave (landmarks)"])
    title_row = sheet.max_row
    sheet.append(["Código", "Aplica a", "Significado"])
    header_row = sheet.max_row
    for row in (
        ("B", "Landmarks pares", "Bilateral visible"),
        ("I", "Landmarks pares", "Solo izquierdo visible"),
        ("D", "Landmarks pares", "Solo derecho visible"),
        ("O", "Todos", "Ocluido o intermitente"),
        ("N", "Todos", "No visible"),
        ("C", "Nariz/centro facial", "Visible"),
    ):
        sheet.append(row)

    title_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="4F6B83")
    thin_border = Border(
        left=Side(style="thin", color="7A8793"),
        right=Side(style="thin", color="7A8793"),
        top=Side(style="thin", color="7A8793"),
        bottom=Side(style="thin", color="7A8793"),
    )
    for cell in sheet[title_row][:3]:
        cell.font = Font(bold=True, color="123B42")
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in sheet[header_row][:3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(
        min_row=title_row,
        max_row=sheet.max_row,
        min_col=1,
        max_col=3,
    ):
        for cell in row:
            cell.border = thin_border
    sheet.merge_cells(
        start_row=title_row,
        start_column=1,
        end_row=title_row,
        end_column=3,
    )


def _paired_landmark_rows(
    summaries: list[dict[str, Any]],
) -> list[list[Any]]:
    grouped: dict[tuple[int, str], dict[str, dict[str, Any]]] = {}
    for summary in summaries:
        key = (
            int(summary["repetition_index"]),
            str(summary["anatomical_group"]),
        )
        grouped.setdefault(key, {})[str(summary["side"])] = summary
    rows: list[list[Any]] = []
    group_order = {
        group: index
        for index, group in enumerate(_LANDMARK_LABELS)
    }
    ordered = sorted(
        grouped.items(),
        key=lambda item: (
            item[0][0],
            group_order.get(item[0][1], len(group_order)),
        ),
    )
    for (repetition, group), sides in ordered:
        left = sides.get("izquierda") or sides.get("central")
        right = sides.get("derecha")
        rows.append(
            [
                repetition,
                _LANDMARK_LABELS.get(group, _display_value(group)),
                _summary_value(left, "mean_visibility"),
                _summary_percentage(left),
                _summary_state(left),
                _summary_value(right, "mean_visibility"),
                _summary_percentage(right),
                _summary_state(right),
                _availability_code(left, right, central="central" in sides),
            ]
        )
    return rows


def _summary_value(
    summary: dict[str, Any] | None,
    key: str,
) -> float | None:
    return round(float(summary[key]), 4) if summary else None


def _summary_percentage(summary: dict[str, Any] | None) -> str | None:
    if not summary:
        return None
    return f"{float(summary['usable_frames_percentage']):.2f} %"


def _summary_state(summary: dict[str, Any] | None) -> str | None:
    return (
        str(_display_value(summary.get("availability")))
        if summary
        else None
    )


def _availability_code(
    left: dict[str, Any] | None,
    right: dict[str, Any] | None,
    *,
    central: bool,
) -> str:
    if central:
        state = left.get("availability") if left else "no_disponible"
        return "C" if state == "visible_estable" else (
            "O" if state == "intermitente" else "N"
        )
    left_state = left.get("availability") if left else "no_disponible"
    right_state = right.get("availability") if right else "no_disponible"
    if left_state == right_state == "visible_estable":
        return "B"
    if left_state == "visible_estable" and right_state == "no_disponible":
        return "I"
    if right_state == "visible_estable" and left_state == "no_disponible":
        return "D"
    if "intermitente" in {left_state, right_state}:
        return "O"
    return "N"


def _with_unit(value: Any, unit: str) -> str | None:
    if value is None:
        return None
    if isinstance(value, float):
        normalized = f"{value:.2f}".rstrip("0").rstrip(".")
        return f"{normalized} {unit}"
    return f"{value} {unit}"


def _technical_cell(value: str) -> Any:
    if value == "":
        return None
    lowered = value.lower()
    if lowered in {"true", "false"}:
        return "Sí" if lowered == "true" else "No"
    side, separator, landmark = lowered.partition("_")
    if separator and side in {"left", "right"} and landmark in _LANDMARK_LABELS:
        return (
            f"{_LANDMARK_LABELS[landmark]} "
            f"{'izquierdo' if side == 'left' else 'derecho'}"
        )
    displayed = _display_value(value)
    if displayed != value:
        return displayed
    try:
        return int(value)
    except ValueError:
        try:
            return float(value)
        except ValueError:
            return value


def _display_value(value: Any) -> Any:
    if value is None or value == "":
        return "No especificado"
    if isinstance(value, bool):
        return "Sí" if value else "No"
    if isinstance(value, list):
        return ", ".join(str(_display_value(item)) for item in value)
    if not isinstance(value, str):
        return value
    if value in _VALUE_LABELS:
        return _VALUE_LABELS[value]
    if value in _PATTERN_NAMES:
        return _PATTERN_NAMES[value]
    try:
        parsed_date = datetime.strptime(value, "%Y-%m-%d")
        return parsed_date.strftime("%d/%m/%Y")
    except ValueError:
        pass
    if "T" in value:
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
            return parsed.strftime("%d/%m/%Y %H:%M:%S")
        except ValueError:
            pass
    return value


def _detected_findings_label(value: Any) -> str:
    if not value:
        return "Ninguna"
    findings = value if isinstance(value, list) else [value]
    return ", ".join(_detected_finding_label(str(item)) for item in findings)


def _detected_finding_label(value: str) -> str:
    repetition, separator, finding = value.partition(":")
    if not separator:
        return str(_display_value(value))
    repetition_number = repetition.removeprefix("repeticion_")
    return f"Repetición {repetition_number}: {_display_value(finding)}"


def _judgment_label(
    classification: str,
    side: str | None,
    pattern_key: str,
) -> str:
    if classification != "presente":
        return _label(classification)
    if pattern_key == "bilateral_asymmetry":
        return "Presente"
    return _label(f"presente_{side or 'sin_direccion'}")


def _label(value: str | None) -> str:
    if value is None:
        return "Pendiente"
    return str(_display_value(value))


def _match_label(value: bool | None) -> str:
    if value is None:
        return "No calculable"
    return "Sí" if value else "No"


def _percent(value: float | None) -> str:
    return f"{value * 100:.1f} %" if value is not None else "N/D"


def _decimal(value: float | None) -> str:
    return f"{value:.3f}" if value is not None else "N/D"


__all__ = [
    "build_case_excel",
    "build_case_pdf",
    "build_technical_data_excel",
]
