"""Tests for researcher-facing Excel and PDF exports."""

from io import BytesIO
from zipfile import ZipFile

from openpyxl import load_workbook

from src.squat.comparison import (
    CaseComparison,
    DatasetPerformance,
    ExpertJudgment,
    FinalReference,
    PatternComparison,
    PerformanceMetrics,
    build_case_comparisons,
)
from src.squat.exports import (
    build_case_excel,
    build_case_pdf,
    build_technical_data_excel,
)


def _comparison() -> CaseComparison:
    rows = build_case_comparisons(
        judgments=[],
        system_decisions=[],
    )
    return CaseComparison(
        case_id="caso_export_001",
        assigned_evaluators=2,
        submitted_evaluations=0,
        patterns=rows,
        ready_for_metrics=False,
    )


def _performance() -> DatasetPerformance:
    empty = PerformanceMetrics(
        scope="general",
        included_pairs=0,
        excluded_inconclusive_pairs=4,
        true_positive=0,
        true_negative=0,
        false_positive=0,
        false_negative=0,
    )
    return DatasetPerformance(
        consolidated_cases=0,
        pending_cases=1,
        overall=empty,
        by_pattern=[],
    )


def _comparison_with_directional_labels() -> CaseComparison:
    pattern = PatternComparison(
        repetition_index=2,
        pattern_key="visible_dynamic_valgus",
        expert_judgments=[
            ExpertJudgment(
                evaluator_id="evaluador-1",
                repetition_index=2,
                pattern_key="visible_dynamic_valgus",
                classification="presente",
                observed_side="izquierda",
            )
        ],
        reference=FinalReference(
            classification="presente",
            observed_side="bilateral",
            method="consenso_guiado",
        ),
        reference_status="consolidada",
        system_classification="presente",
        system_side="derecha",
        system_label="presente_derecha",
        exact_match=False,
        binary_outcome="TP",
    )
    return CaseComparison(
        case_id="caso_export_001",
        assigned_evaluators=1,
        submitted_evaluations=1,
        patterns=[pattern],
        ready_for_metrics=True,
    )


def test_excel_contains_instruments_and_analysis_sheets() -> None:
    content = build_case_excel(
        case_record={
            "created_at": "2026-07-29T08:45:46.624700Z",
            "registration": {
                "case": {
                    "case_id": "caso_export_001",
                    "participant_code": "P-001",
                    "participant_age": 28,
                    "participant_sex": "femenino",
                    "view": "anterior",
                    "plane": "frontal",
                    "load_condition": "sin_carga_externa",
                },
                "video": {
                    "path": "video.mp4",
                    "width_px": 1080,
                    "height_px": 1920,
                    "fps": 30,
                },
                "ready_for_pose": True,
            },
        },
        case_report={
            "case_id": "caso_export_001",
            "status": "analisis_completo",
            "pose": {
                "total_frames": 120,
                "processed_frames": 118,
                "valid_frames": 110,
                "valid_frames_percentage": 91.67,
                "processed_frames_percentage": 98.33,
                "mean_detected_keypoints": 12.4,
            },
            "findings": {
                "detected_findings": [
                    "repeticion_2:valgo_dinamico_visible",
                    "repeticion_2:asimetria_bilateral_observable",
                ],
                "decisions": [],
                "ruleset_version": "test",
            },
        },
        comparison=_comparison_with_directional_labels(),
        performance=_performance(),
        landmark_visibility=[
            {
                "repetition_index": 1,
                "landmark": "left_hip",
                "anatomical_group": "hip",
                "side": "izquierda",
                "mean_visibility": 0.92,
                "usable_frames_percentage": 100,
                "availability": "visible_estable",
            },
            {
                "repetition_index": 1,
                "landmark": "right_hip",
                "anatomical_group": "hip",
                "side": "derecha",
                "mean_visibility": 0.88,
                "usable_frames_percentage": 95,
                "availability": "visible_estable",
            },
        ],
    )

    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == [
        "Instrumento 1",
        "Instrumento 2",
        "Instrumento 3",
        "Matriz de análisis",
        "Métricas",
    ]
    instrument_1 = workbook["Instrumento 1"]
    assert instrument_1["A2"].value == "Código del video"
    assert instrument_1["B2"].value == "caso_export_001"
    assert instrument_1["B4"].value == 28
    assert instrument_1["B5"].value == "Femenino"
    assert instrument_1["B14"].value == "1080 × 1920 px"
    assert instrument_1["B15"].value == "30 fps"
    assert not any(
        str(cell.value).startswith("registration.")
        for row in instrument_1.iter_rows()
        for cell in row
        if cell.value
    )
    assert any(
        cell.value == "Cadera"
        for row in instrument_1.iter_rows()
        for cell in row
    )
    legend_title = next(
        cell
        for row in instrument_1.iter_rows()
        for cell in row
        if cell.value == "Codificación de puntos anatómicos clave (landmarks)"
    )
    assert all(
        (
            instrument_1.cell(
                row=legend_title.row - offset,
                column=column,
            ).value
            is None
        )
        for offset in (1, 2)
        for column in range(1, 4)
    )
    assert instrument_1.cell(row=legend_title.row + 2, column=1).value == "B"
    assert (
        instrument_1.cell(row=legend_title.row + 2, column=3).value
        == "Bilateral visible"
    )
    assert legend_title.fill.fgColor.rgb == "00D9EAF7"
    legend_header = instrument_1.cell(row=legend_title.row + 1, column=1)
    assert legend_header.fill.fgColor.rgb == "004F6B83"
    assert legend_header.font.color.rgb == "00FFFFFF"
    assert legend_header.border.bottom.style == "thin"
    assert (
        instrument_1.cell(row=legend_title.row + 2, column=3).border.right.style
        == "thin"
    )

    instrument_2 = workbook["Instrumento 2"]
    headers = [cell.value for cell in instrument_2[1]]
    processed_column = headers.index("Fotogramas procesados correctamente") + 1
    findings_column = headers.index("Compensaciones detectadas") + 1
    assert instrument_2.cell(row=2, column=processed_column).value == 118
    assert instrument_2.cell(row=2, column=findings_column).value == (
        "Repetición 2: Valgo dinámico visible, "
        "Repetición 2: Asimetría bilateral observable"
    )
    assert workbook["Instrumento 3"]["A2"].value == "caso_export_001"
    assert workbook["Instrumento 3"]["D2"].value == "Presente, izquierda"
    assert workbook["Instrumento 3"]["G2"].value == "Presente, derecha"
    assert workbook["Instrumento 3"]["H2"].value == "Presente, bilateral"
    assert workbook["Matriz de análisis"]["C2"].value == "Presente, bilateral"
    assert workbook["Matriz de análisis"]["D2"].value == "Presente, derecha"


def test_pdf_export_has_a_valid_header() -> None:
    content = build_case_pdf(
        case_report={
            "case_id": "caso_export_001",
            "pipeline_version": "test",
        },
        comparison=_comparison(),
        performance=_performance(),
    )

    assert content.startswith(b"%PDF")
    assert len(content) > 1_000
    assert content.count(b"/Type /Page") >= 3


def test_technical_excel_normalizes_csv_without_mutating_source() -> None:
    source = (
        b"frame_index,timestamp_seconds,pose_detected,valid_for_analysis,"
        b"detected_keypoints,minimum_critical_visibility,invalid_reason\n"
        b"0,0.0,True,True,13,0.91,\n"
    )

    content = build_technical_data_excel(
        artifacts={"frame_quality.csv": source}
    )

    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Calidad fotogramas"]
    assert sheet["A1"].value == "N.° de fotograma"
    assert sheet["C1"].value == "Pose detectada"
    assert sheet["C2"].value == "Sí"
    assert sheet["E2"].value == 13
    assert len(sheet.tables) == 0
    assert sheet.auto_filter.ref == "A1:G2"
    with ZipFile(BytesIO(content)) as archive:
        assert not any(
            name.startswith("xl/tables/") for name in archive.namelist()
        )
    assert source.startswith(b"frame_index,timestamp_seconds")
